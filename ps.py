import jieba.posseg as pseg

import jieba.analyse
import xlwt
import openpyxl
from gensim import corpora, models, similarities
import re

#停词函数
def StopWordsList(filepath):
    wlst = [w.strip() for w in open(filepath, 'r', encoding='utf8').readlines()]
    return wlst

def str_to_hex(s):
    return ''.join([hex(ord(c)).replace('0x', '') for c in s])

# jieba分词
def seg_sentence(sentence, stop_words):
    stop_flag = ['x', 'c', 'u', 'd', 'p', 't', 'uj', 'f', 'r']
    sentence_seged = pseg.cut(sentence)
    outstr = []
    for word, flag in sentence_seged:
        if word not in stop_words and flag not in stop_flag:
            outstr.append(word)
    return outstr

if __name__ == '__main__':
    #1 这些是jieba分词的自定义词典，软件老王这里添加的格式行业术语，格式就是文档，一列一个词一行就行了，
    # 这个几个词典软件老王就不上传了，可注释掉。
    jieba.load_userdict("g1.txt")
    jieba.load_userdict("g2.txt")
    jieba.load_userdict("g3.txt")

    #2 停用词，简单理解就是这次词不分割，这个软件老王找的网上通用的，会提交下。
    spPath = 'stop.txt'
    stop_words = StopWordsList(spPath)

    #3 excel处理
    wbk = xlwt.Workbook(encoding='ascii')
    sheet = wbk.add_sheet("软件老王sheet")  # sheet名称
    sheet.write(0, 0, '表头-软件老王1')
    sheet.write(0, 1, '表头-软件老王2')
    sheet.write(0, 2, '导航-链接到明细sheet表')
    wb = openpyxl.load_workbook('软件老王-source.xlsx')
    ws = wb.active
    col = ws['B']
    # 4 相似性处理
    rcount = 1
    texts = []
    orig_txt = []
    key_list = []
    name_list = []
    sheet_list = []

    for cell in col:
        if cell.value is None:
            continue
        if not isinstance(cell.value, str):
            continue
        item = cell.value.strip('\n\r').split('\t')  # 制表格切分
        string = item[0]
        if string is None or len(string) == 0:
            continue
        else:
            textstr = seg_sentence(string, stop_words)
            texts.append(textstr)
            orig_txt.append(string)
    dictionary = corpora.Dictionary(texts)
    feature_cnt = len(dictionary.token2id.keys())
    corpus = [dictionary.doc2bow(text) for text in texts]
    tfidf = models.LsiModel(corpus)
    index = similarities.SparseMatrixSimilarity(tfidf[corpus], num_features=feature_cnt)
    result_lt = []
    word_dict = {}
    count =0
    for keyword in orig_txt:
        count = count+1
        print('开始执行，第'+ str(count)+'行')
        if keyword in result_lt or keyword is None or len(keyword) == 0:
            continue
        kw_vector = dictionary.doc2bow(seg_sentence(keyword, stop_words))
        sim = index[tfidf[kw_vector]]
        result_list = []
        for i in range(len(sim)):
            if sim[i] > 0.5:
                if orig_txt[i] in result_lt and orig_txt[i] not in result_list:
                    continue
                result_list.append(orig_txt[i])
                result_lt.append(orig_txt[i])
        if len(result_list) >0:
            word_dict[keyword] = len(result_list)
        if len(result_list) >= 1:
            sname = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", "", keyword[0:10])+ '_'\
                    + str(len(result_list)+ len(str_to_hex(keyword))) + str_to_hex(keyword)[-5:]
            sheet_t = wbk.add_sheet(sname)  # Excel单元格名字
            for i in range(len(result_list)):
                sheet_t.write(i, 0, label=result_list[i])

    #5 按照热度排序 -软件老王
    with open("rjlw.txt", 'w', encoding='utf-8') as wf2:
        orderList = list(word_dict.values())
        orderList.sort(reverse=True)
        count = len(orderList)
        for i in range(count):
            for key in word_dict:
                if word_dict[key] == orderList[i]:
                    key_list.append(key)
                    word_dict[key] = 0
        wf2.truncate()
    #6 写入目标excel
    for i in range(len(key_list)):
        sheet.write(i+rcount, 0, label=key_list[i])
        sheet.write(i+rcount, 1, label=orderList[i])
        if orderList[i] >= 1:
            shname = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", "", key_list[i][0:10]) \
                     + '_'+ str(orderList[i]+ len(str_to_hex(key_list[i])))+ str_to_hex(key_list[i])[-5:]
            link = 'HYPERLINK("#%s!A1";"%s")' % (shname, shname)
            sheet.write(i+rcount, 2, xlwt.Formula(link))
    rcount = rcount + len(key_list)
    key_list = []
    orderList = []
    texts = []
    orig_txt = []
    wbk.save('软件老王-target.xls')